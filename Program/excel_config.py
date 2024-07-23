from openpyxl import *
from openpyxl.chart import PieChart3D, Reference

while True:


    choice = input("Enter in your choice (1, 2, or 3): ")

    if choice == "1":

        name = input("Enter in the name of your Excel workbook (including extension): ")

        data = [

            ["Crisps", "Sold"],
            ["Salty", 100],
            ["Onion", 94],
            ["Chilli", 88],
            ["Chicken", 54],
            ["Bacon", 21],
        ]

        book = Workbook()

        sheet = book.active

        for row in data:

            sheet.append(row)

        chart = PieChart3D()

        labels = Reference(sheet, min_col=1, min_row=2, max_row=5)

        data2 = Reference(sheet, min_col=2, min_row=1, max_row=5)

        chart.add_data(data2, titles_from_data=True)

        chart.set_categories(labels)

        chart.title = "Most Popular Crisps Flavor"

        sheet.add_chart(chart, "A8")

        book.save(name)

    elif choice == "2":

        name = input("Enter in the name of your Excel workbook (including extension): ")

        num1 = input("Enter in a number: ")

        num2 = input("Enter in another number: ")

        book = Workbook()

        sheet = book.active

        sheet["A1"] = num1

        sheet["B1"] = num2

        sheet["C1"] = "SUM(A1, B1)"

        sheet["D1"] = "IF(A1>B1, \"A1 is greater than B1\", \"B1 is greater than A1\")"

        book.save(name)

    elif choice == "3": 

        name = input("Enter in the name of your Excel workbook (including extension): ")

        sheet_name1 = input("Enter in the name of the first sheet: ")

        sheet_name2 = input("Enter in the name of the second sheet: ")

        def sheet_creator(file):
        
            book = Workbook()
            
            sheet = book.active
            
            sheet.title = sheet_name1

            sheet2 = book.create_sheet(title=sheet_name2)

            book.save(file)


        if __name__ == '__main__':

            sheet_creator(name)

    else:

        print("Invalid choice!")
    
    input("Do you want to select another choice? (y/n): ")

    if choice == "n" or choice == "N":

        break